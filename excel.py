#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author:Newton_apple time:2018/2/19
# 读写excel

import numpy as np
from openpyxl import load_workbook

class Excel:
	"""excel表格存储读取矩阵"""
	def __init__(self, filename='matrix.xlsx'):
		self.wb = load_workbook(filename)
		self.sheet = self.wb['Sheet1']

	def get_matrix(self):
		"""读取excle里的数据，并生成矩阵，返回"""
		s = ''
		for i in tuple(self.sheet.rows):
			for cell in i:
				s = s + str(cell.value) + ' '
			s = s + ';'
		s = s[:-2]  # 去掉最后面那个分号
		mat_rix = np.matrix(s)
		return mat_rix

if __name__ == '__main__':
	ex = Excel()
	nu = ex.get_matrix()
	print(nu)
	wb = load_workbook('matrix.xlsx')
	print(wb.sheetnames)
	sheet = wb['Sheet1']   # 弃用了wb.get_sheet_by_name('Sheet1')方法
	# for i in sheet['2']:
	# 	print(i.value, end=' ')
	print(tuple(sheet.columns))
	s = ''
	for i in tuple(sheet.rows):
		for cell in i:
			s = s + str(cell.value) + ' '
		s = s + ';'
	s = s[:-2]          # 去掉最后面那个分号
	print(s)

